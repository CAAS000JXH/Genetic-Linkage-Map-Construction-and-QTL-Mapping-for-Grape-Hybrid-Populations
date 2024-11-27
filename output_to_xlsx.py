import sys
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def apply_formatting(sheet, parent_type):
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    red_fill = PatternFill(start_color="FFC0C0", end_color="FFC0C0", fill_type="solid")

    # 设置单元格颜色
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=sheet.max_column):
        for cell in row:
            if cell.value in ["nn", "ll"]:
                cell.fill = yellow_fill
            elif cell.value in ["np", "lm"]:
                cell.fill = red_fill
    
    # 设置列宽
    for col in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 2.5

def modify_header(sheet, header_file):
    with open(header_file, 'r') as f:
        new_headers = f.read().splitlines()

    # 替换表头
    for col_index, new_header in enumerate(new_headers, start=3):
        sheet.cell(row=1, column=col_index, value=new_header)

def add_segregation_type(sheet, parent_type):
    segregation_type = "<nnxnp>" if parent_type == "male" else "<lmxll>"
    sheet.insert_cols(2)  # 在第二列插入一列
    sheet.cell(row=1, column=2, value="Segregation Type")  # 添加列标题
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=2, value=segregation_type)

def process_chromosome(chrom_id, parent_type, output_wb, header_file):
    input_file = f'{chrom_id}.bins.genotype.{parent_type}.correction.csv'
    with open(input_file, mode='r') as infile:
        reader = csv.reader(infile, delimiter=';')
        data = list(reader)

    # 创建新的工作表并命名为染色体ID
    sheet = output_wb.create_sheet(title=chrom_id)
    
    # 将CSV数据写入工作表
    for row_index, row_data in enumerate(data):
        for col_index, cell_value in enumerate(row_data):
            sheet.cell(row=row_index+1, column=col_index+1, value=cell_value)
    
    # 修改表头
    modify_header(sheet, header_file)
    
    # 插入分离类型列
    add_segregation_type(sheet, parent_type)
    
    # 应用格式
    apply_formatting(sheet, parent_type)

def main(chromosome_file, parent_type, output_file, header_file):
    # 读取染色体ID文件
    with open(chromosome_file, 'r') as f:
        chrom_ids = f.read().splitlines()

    if parent_type not in ["male", "female"]:
        raise ValueError("亲本类型只能是 'male' 或 'female'")
    
    # 创建Excel工作簿
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # 删除默认创建的工作表

    # 处理每个染色体ID
    for chrom_id in chrom_ids:
        process_chromosome(chrom_id, parent_type, output_wb, header_file)
    
    # 保存输出文件
    output_wb.save(output_file)
    print(f"结果已保存至: {output_file}")

if __name__ == '__main__':
    if len(sys.argv) != 5:
        print("Usage: python script.py <chromosome_file> <parent_type> <output_file> <header_file>")
    else:
        chromosome_file = sys.argv[1]
        parent_type = sys.argv[2]
        output_file = sys.argv[3]
        header_file = sys.argv[4]
        main(chromosome_file, parent_type, output_file, header_file)
